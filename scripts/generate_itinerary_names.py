"""
Itinerary name generator (Python) — Bali v2.

Goal: ≤24 chars, unique within destination (target: 18,857+ distinct).
Approach:
  - Expanded word pools (~3× prior version)
  - Compound templates: {adj} {noun}, {adj} {city} {noun}, {verb} the {noun}, etc.
  - Edge-case guards: skip null cities, normalise casing, dedupe same-city routes
  - Human-tone fallback: "Bali Chapter N" / "Bali Edition N" — never SKU-style
"""

import json
import re
from collections import Counter

# ─── Word banks ───
BALI = {
    "nouns": [
        # Concrete scenic (plural only)
        "Paddies", "Swings", "Temples", "Beaches", "Reefs", "Cliffs", "Caves",
        "Waterfalls", "Volcanoes", "Sunsets", "Sunrises", "Villas", "Jungles",
        "Terraces", "Moons", "Bonfires", "Lanterns", "Pools", "Shores",
        "Lagoons", "Mists", "Spices", "Courtyards", "Breezes", "Shrines", "Markets",
        "Rituals", "Harvests", "Cascades", "Gardens", "Bridges", "Stars", "Tides",
        # Experience
        "Trails", "Routes", "Voyages", "Journeys", "Escapes", "Retreats", "Chapters",
        "Mornings", "Evenings", "Afternoons", "Nights", "Dreams", "Tales", "Stories",
        # Cultural
        "Dances", "Songs", "Feasts", "Crafts", "Silks", "Offerings",
        # Food / dining
        "Dinners", "Breakfasts", "Picnics", "Flavours",
        # Aquatic
        "Waves", "Coves", "Harbours",
        # Floral / nature
        "Orchids", "Canopies", "Fields", "Sands",
    ],
    "adjs": [
        # Atmospheric
        "Sacred", "Jungle", "Island", "Slow", "Hidden", "Secret", "Wild", "Laid-back",
        "Royal", "Barefoot", "Lush", "Quiet", "Sunlit", "Soulful", "Tropical",
        "Coastal", "Golden", "Starlit", "Whispered", "Dreamy", "Untamed", "Storied",
        "Breezy", "Moonlit", "Sunkissed", "Salty", "Volcanic",
        # Emotional
        "Blissful", "Tender", "Intimate", "Quiet", "Mellow", "Serene", "Private",
        "Timeless", "Languid", "Easy", "Effortless",
        # Temporal
        "Eternal", "Boundless", "Endless",
        # Travel-y
        "Offbeat", "Wanderer", "Roaming", "Winding", "Slow-burn",
    ],
    "verbs": [
        "Chase", "Swing", "Snorkel", "Sip", "Wander", "Roam", "Dive", "Ride",
        "Hike", "Unwind", "Linger", "Drift", "Feast", "Stroll", "Rest", "Bike",
        "Sail", "Climb", "Chase", "Explore", "Savour", "Whisper", "Breathe",
        "Float", "Ramble", "Tread", "Follow", "Seek", "Trace", "Dance",
    ],
    "feelings": [
        "Bliss", "Calm", "Serenity", "Wanderlust", "Joy", "Magic", "Escape",
        "Togetherness", "Stillness", "Wonder", "Peace", "Awe", "Dreaming",
        "Afterglow", "Enchantment", "Warmth", "Glow", "Hush", "Daydream",
        "Romance", "Harmony", "Ease",
    ],
    "love_words": ["Romance", "Love", "Honeymoon", "Us-time", "Moments", "Two"],
    "scales": {
        "short": ["Short", "Quick", "Weekend", "Mini", "Pocket"],
        "long": ["Grand", "All of", "Best of", "Endless", "Full"],
    },
    # Tag-preferred sub-pools
    "adventure_nouns": ["Swings", "Volcanoes", "Cliffs", "Jungles", "Waterfalls", "Trails", "Caves", "Waves", "Canopies"],
    "romance_nouns": ["Sunsets", "Moons", "Villas", "Bonfires", "Pools", "Lanterns", "Mornings", "Evenings", "Picnics", "Dinners"],
    "culture_nouns": ["Temples", "Terraces", "Shrines", "Markets", "Rituals", "Spices", "Silks", "Offerings", "Dances"],
    "water_nouns": ["Reefs", "Shores", "Lagoons", "Beaches", "Sunrises", "Coves", "Harbours", "Waves", "Tides", "Sands"],
    "nature_nouns": ["Paddies", "Terraces", "Mists", "Gardens", "Waterfalls", "Orchids", "Fields"],
    "adventure_verbs": ["Chase", "Climb", "Dive", "Hike", "Ride", "Swing", "Surf", "Trace", "Seek"],
    "romance_verbs": ["Sip", "Unwind", "Linger", "Drift", "Whisper", "Rest", "Dance", "Savour"],
}

# ─── Keyword-to-tag map for activities ───
KEYWORD_TAGS = {
    "adventure": ["swing", "atv", "trek", "dive", "snorkel", "volcano", "hike", "climb", "surf", "rafting", "jeep", "cycling"],
    "romance": ["candle", "sunset", "romantic", "dinner", "private", "couple", "honeymoon", "spa"],
    "culture": ["temple", "heritage", "village", "dance", "cooking", "market", "palace", "ceremony", "silk"],
    "water": ["snorkel", "boat", "cruise", "beach", "reef", "dive", "island", "lagoon", "fishing"],
    "nature": ["rice", "waterfall", "jungle", "forest", "mountain", "sunrise", "spring", "garden"],
    "wellness": ["spa", "yoga", "wellness", "massage", "retreat"],
}

def extract_tags(activities):
    if not activities: return set()
    tags = set()
    blob = " ".join((a.get("name") or "") for a in activities).lower() if isinstance(activities, list) else str(activities).lower()
    for tag, kws in KEYWORD_TAGS.items():
        if any(kw in blob for kw in kws):
            tags.add(tag)
    return tags

# ─── Hashing ───
def djb2(s):
    h = 5381
    for ch in s:
        h = ((h << 5) + h + ord(ch)) & 0xFFFFFFFF
    return h

# Primes chosen so different slots rotate through pools independently;
# dramatically widens the effective combinatorial namespace for 3-slot
# templates (≈ product of pool sizes rather than sum).
_SLOT_PRIMES = [2654435761, 40503289, 67867967, 15485863, 982451653, 1300007, 3000017, 900091, 100003, 10007]

def _pick_factory(bank, seed):
    def pick(category, offset=0, sub=None):
        pool = bank[category][sub] if sub else bank[category]
        prime = _SLOT_PRIMES[offset % len(_SLOT_PRIMES)]
        return pool[(seed * prime) % len(pool)]
    return pick

# Helper: get an int index for arbitrary-length collections without a named bank.
def _raw_idx(seed, n, offset=0):
    return (seed + offset) % n

# ─── Helpers ───
_CITY_ALIASES = {
    "gili trawangan": "Gili T",
    "nusa lembongan": "Lembongan",
    "labuan bajo": "Komodo",
    "karangasem": "Karangasem",
    "canggu": "Canggu",  # normalise casing
}

def _clean_city(c):
    if not c: return None
    s = str(c).strip()
    if not s or s.lower() == "null": return None
    # alias + title-case
    return _CITY_ALIASES.get(s.lower(), s.title() if s.isupper() else s)

def _normalise_cities(route):
    """Deduplicate same-city repeats; drop nulls; apply alias/casing."""
    seen = []
    for seg in route:
        c = _clean_city(seg.get("city"))
        if c and (not seen or seen[-1] != c):
            seen.append(c)
    return seen

# ─── Tag-aware noun/verb selection ───
def _noun_pool_for_tags(tags, salt):
    pref = []
    if "romance" in tags: pref.append("romance_nouns")
    if "adventure" in tags: pref.append("adventure_nouns")
    if "culture" in tags: pref.append("culture_nouns")
    if "water" in tags: pref.append("water_nouns")
    if "nature" in tags: pref.append("nature_nouns")
    return pref[salt % len(pref)] if pref else "nouns"

def _verb_pool_for_tags(tags, salt):
    pref = []
    if "adventure" in tags: pref.append("adventure_verbs")
    if "romance" in tags: pref.append("romance_verbs")
    return pref[salt % len(pref)] if pref else "verbs"

# ─── Templates ───
def _t_two_cities(ctx):
    c = ctx["cities"]
    if len(c) < 2: return None
    s = f"{c[0]} & {c[-1]}"
    return s if len(s) <= 24 else None

def _t_cities_to(ctx):
    c = ctx["cities"]
    if len(c) < 2: return None
    s = f"{c[0]} to {c[-1]}"
    return s if len(s) <= 24 else None

def _t_adj_city(ctx):
    c = ctx["cities"]
    if not c: return None
    s = f"{ctx['pick']('adjs')} {c[0]}"
    return s if len(s) <= 24 else None

def _t_city_possessive(ctx):
    c = ctx["cities"]
    if not c: return None
    np = _noun_pool_for_tags(ctx["tags"], 3)
    s = f"{c[0]}'s {ctx['pick'](np, 2)}"
    return s if len(s) <= 24 else None

def _t_noun_and_noun(ctx):
    p1 = _noun_pool_for_tags(ctx["tags"], 0)
    p2 = _noun_pool_for_tags(ctx["tags"], 4)
    n1 = ctx['pick'](p1)
    n2 = ctx['pick'](p2, 1)
    if n1 == n2: return None
    s = f"{n1} & {n2}"
    return s if len(s) <= 24 else None

def _t_verbs(ctx):
    vp = _verb_pool_for_tags(ctx["tags"], 0)
    v1 = ctx['pick'](vp)
    v2 = ctx['pick'](vp, 1)
    v3 = ctx['pick'](vp, 2)
    if len({v1, v2, v3}) < 3: return None
    s = f"{v1}, {v2} & {v3}"
    return s if len(s) <= 24 else None

def _t_adj_dest(ctx):
    s = f"{ctx['pick']('adjs')} {ctx['dest']}"
    return s if len(s) <= 24 else None

def _t_feeling_in_city(ctx):
    where = ctx["cities"][0] if ctx["cities"] else ctx["dest"]
    s = f"{ctx['pick']('feelings')} in {where}"
    return s if len(s) <= 24 else None

def _t_nights_of(ctx):
    n = ctx["nights"]
    if not n: return None
    s = f"{n}N of {ctx['pick']('feelings')}"
    return s if len(s) <= 24 else None

def _t_scale_dest(ctx):
    n = ctx["nights"]
    if not n: return None
    key = "short" if n <= 5 else ("long" if n >= 10 else None)
    if not key: return None
    scale = ctx["pick"]("scales", 0, key)
    s = f"{scale} {ctx['dest']}"
    return s if len(s) <= 24 else None

def _t_two_hearts(ctx):
    s = f"Two Hearts, Two {ctx['pick']('nouns')}"
    return s if len(s) <= 24 else None

# High-entropy compounds
def _t_adj_noun(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 3)
    s = f"{ctx['pick']('adjs')} {ctx['pick'](np, 5)}"
    return s if len(s) <= 24 else None

def _t_city_noun(ctx):
    c = ctx["cities"]
    if not c: return None
    np = _noun_pool_for_tags(ctx["tags"], 2)
    s = f"{c[0]} {ctx['pick'](np, 6)}"
    return s if len(s) <= 24 else None

# Semantically valid {verb} the {noun} pairs. Keeps names grammatical
# ("Dive the Reefs" ✓, "Swing the Waves" ✗).
_VERB_NOUN_PAIRS = []
_VN_MAP = {
    "Dive":    ["Reefs", "Coves", "Lagoons", "Waves"],
    "Hike":    ["Trails", "Jungles", "Volcanoes", "Cliffs", "Canopies"],
    "Climb":   ["Cliffs", "Volcanoes", "Trails"],
    "Surf":    ["Waves", "Shores"],
    "Chase":   ["Sunsets", "Sunrises", "Waves", "Tides", "Stars", "Moons"],
    "Stroll":  ["Markets", "Gardens", "Shores", "Beaches", "Courtyards", "Paddies"],
    "Wander":  ["Markets", "Jungles", "Trails", "Shores", "Paddies", "Terraces"],
    "Roam":    ["Trails", "Shores", "Jungles", "Fields", "Gardens"],
    "Ride":    ["Waves", "Trails"],
    "Sail":    ["Coves", "Lagoons", "Shores", "Tides"],
    "Explore": ["Reefs", "Jungles", "Temples", "Markets", "Caves", "Waterfalls", "Trails", "Shrines"],
    "Seek":    ["Waterfalls", "Sunrises", "Caves", "Shrines"],
    "Trace":   ["Trails", "Routes"],
    "Tread":   ["Trails"],
    "Follow":  ["Trails", "Tides", "Sunsets"],
    "Swing":   [],  # no valid direct-object pairing
    "Sip":     [],
    "Snorkel": ["Reefs", "Lagoons", "Coves"],
    "Drift":   ["Tides"],
    "Ramble":  ["Trails", "Markets"],
    "Savour":  ["Sunsets", "Moons", "Evenings", "Dinners"],
    "Bike":    ["Trails", "Paddies"],
}
for v, ns in _VN_MAP.items():
    for n in ns: _VERB_NOUN_PAIRS.append((v, n))

def _t_verb_the_noun(ctx):
    if not _VERB_NOUN_PAIRS: return None
    v, n = _VERB_NOUN_PAIRS[ctx["seed"] % len(_VERB_NOUN_PAIRS)]
    s = f"{v} the {n}"
    return s if len(s) <= 24 else None

def _t_adj_city_noun(ctx):
    c = ctx["cities"]
    if not c: return None
    np = _noun_pool_for_tags(ctx["tags"], 7)
    s = f"{ctx['pick']('adjs')} {c[0]} {ctx['pick'](np, 8)}"
    return s if len(s) <= 24 else None

def _t_city_adj_noun(ctx):
    c = ctx["cities"]
    if not c: return None
    np = _noun_pool_for_tags(ctx["tags"], 9)
    s = f"{c[0]} {ctx['pick']('adjs', 1)} {ctx['pick'](np, 10)}"
    return s if len(s) <= 24 else None

def _t_adj_adj_noun(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 2)
    s = f"{ctx['pick']('adjs')} {ctx['pick']('adjs', 3)} {ctx['pick'](np, 4)}"
    return s if len(s) <= 24 else None

def _t_noun_of_city(ctx):
    c = ctx["cities"]
    if not c: return None
    np = _noun_pool_for_tags(ctx["tags"], 4)
    s = f"{ctx['pick'](np)} of {c[0]}"
    return s if len(s) <= 24 else None

def _t_a_noun_in_city(ctx):
    c = ctx["cities"]
    if not c: return None
    np = _noun_pool_for_tags(ctx["tags"], 5)
    s = f"A {ctx['pick'](np)} in {c[0]}"
    return s if len(s) <= 24 else None

def _t_noun_of_feeling(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 6)
    s = f"{ctx['pick'](np)} of {ctx['pick']('feelings', 2)}"
    return s if len(s) <= 24 else None

def _t_nights_noun_trail(ctx):
    n = ctx["nights"]
    if not n: return None
    np = _noun_pool_for_tags(ctx["tags"], 8)
    s = f"{n}N {ctx['pick'](np)}"
    return s if len(s) <= 24 else None

# High-entropy compounds (no city):
def _t_adj_noun_and_noun(ctx):
    np1 = _noun_pool_for_tags(ctx["tags"], 3)
    np2 = _noun_pool_for_tags(ctx["tags"], 5)
    n1 = ctx['pick'](np1, 2)
    n2 = ctx['pick'](np2, 4)
    if n1 == n2: return None
    s = f"{ctx['pick']('adjs')} {n1} & {n2}"
    return s if len(s) <= 24 else None

def _t_noun_noun_feeling(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 2)
    n1 = ctx['pick'](np)
    n2 = ctx['pick'](np, 3)
    if n1 == n2: return None
    s = f"{n1}, {n2} & {ctx['pick']('feelings', 1)}"
    return s if len(s) <= 24 else None

def _t_the_adj_noun(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 1)
    s = f"The {ctx['pick']('adjs')} {ctx['pick'](np, 4)}"
    return s if len(s) <= 24 else None

def _t_feeling_of_noun(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 3)
    s = f"{ctx['pick']('feelings')} of {ctx['pick'](np, 2)}"
    return s if len(s) <= 24 else None

def _t_nights_adj_dest(ctx):
    n = ctx["nights"]
    if not n: return None
    s = f"{n}N {ctx['pick']('adjs')} {ctx['dest']}"
    return s if len(s) <= 24 else None

def _t_adj_noun_of_dest(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 5)
    s = f"{ctx['pick']('adjs')} {ctx['pick'](np, 2)} of {ctx['dest']}"
    return s if len(s) <= 24 else None

# --- New high-namespace templates ---
def _t_adj_noun_of_feeling(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 6)
    s = f"{ctx['pick']('adjs')} {ctx['pick'](np, 2)} of {ctx['pick']('feelings', 4)}"
    return s if len(s) <= 24 else None

def _t_nights_of_adj_noun(ctx):
    n = ctx["nights"]
    if not n: return None
    np = _noun_pool_for_tags(ctx["tags"], 3)
    s = f"{n}N of {ctx['pick']('adjs')} {ctx['pick'](np, 5)}"
    return s if len(s) <= 24 else None

def _t_verb_the_adj_noun(ctx):
    # Use curated verb-noun pairs with an adjective inserted
    if not _VERB_NOUN_PAIRS: return None
    v, n = _VERB_NOUN_PAIRS[ctx["seed"] % len(_VERB_NOUN_PAIRS)]
    a = ctx["pick"]("adjs", 3)
    s = f"{v} the {a} {n}"
    return s if len(s) <= 24 else None

def _t_noun_feeling_noun(ctx):
    np1 = _noun_pool_for_tags(ctx["tags"], 1)
    np2 = _noun_pool_for_tags(ctx["tags"], 7)
    s = f"{ctx['pick'](np1)} & {ctx['pick']('feelings', 2)} {ctx['pick'](np2, 3)}"
    return s if len(s) <= 24 else None

def _t_feeling_among_nouns(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 4)
    s = f"{ctx['pick']('feelings')} Among {ctx['pick'](np, 2)}"
    return s if len(s) <= 24 else None

def _t_a_feeling_journey(ctx):
    journey_pool = ["Journey", "Escape", "Getaway", "Trail", "Tale", "Chapter", "Story"]
    j = journey_pool[ctx["seed"] % len(journey_pool)]
    s = f"A {ctx['pick']('feelings')} {j}"
    return s if len(s) <= 24 else None

def _t_dest_of_feeling(ctx):
    s = f"{ctx['dest']} of {ctx['pick']('feelings')}"
    return s if len(s) <= 24 else None

def _t_nights_adj_noun(ctx):
    n = ctx["nights"]
    if not n: return None
    np = _noun_pool_for_tags(ctx["tags"], 4)
    s = f"{n}N of {ctx['pick']('adjs')} {ctx['pick'](np, 7)}"
    return s if len(s) <= 24 else None

def _t_three_nouns(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 2)
    n1 = ctx['pick'](np)
    n2 = ctx['pick'](np, 3)
    n3 = ctx['pick'](np, 5)
    if len({n1, n2, n3}) < 3: return None
    s = f"{n1}, {n2} & {n3}"
    return s if len(s) <= 24 else None

def _t_adj_feeling_noun(ctx):
    np = _noun_pool_for_tags(ctx["tags"], 6)
    # e.g. "Endless Bliss Mornings"  — adjective strengthens the compound
    s = f"{ctx['pick']('adjs')} {ctx['pick']('feelings', 2)} {ctx['pick'](np, 4)}"
    return s if len(s) <= 24 else None

def _t_city_and_beyond(ctx):
    c = ctx["cities"]
    if not c: return None
    s = f"{c[0]} & Beyond"
    return s if len(s) <= 24 else None

def _t_verb_in_city(ctx):
    c = ctx["cities"]
    if not c: return None
    vp = _verb_pool_for_tags(ctx["tags"], 0)
    s = f"{ctx['pick'](vp)} in {c[0]}"
    return s if len(s) <= 24 else None

# Love-tinted (gated)
def _t_love_among(ctx):
    s = f"{ctx['pick']('love_words')} Among {ctx['pick']('nouns')}"
    return s if len(s) <= 24 else None

def _t_love_in_dest(ctx):
    s = f"{ctx['pick']('love_words')} in {ctx['dest']}"
    return s if len(s) <= 24 else None

def _t_city_love(ctx):
    c = ctx["cities"]
    if not c: return None
    s = f"{c[0]} {ctx['pick']('love_words')}"
    return s if len(s) <= 24 else None

# Templates that DON'T reference cities — used whenever the itinerary has
# activity signal (the common case). These derive from activity-tagged nouns,
# feelings, verbs, and the destination name.
ACTIVITY_TEMPLATES = [
    _t_noun_and_noun, _t_verbs, _t_adj_dest, _t_nights_of, _t_scale_dest,
    _t_two_hearts, _t_adj_noun, _t_verb_the_noun, _t_noun_of_feeling,
    _t_nights_noun_trail,
    _t_adj_noun_and_noun, _t_noun_noun_feeling, _t_the_adj_noun,
    _t_feeling_of_noun, _t_nights_adj_dest, _t_adj_noun_of_dest,
    _t_adj_noun_of_feeling, _t_nights_of_adj_noun, _t_verb_the_adj_noun,
    _t_feeling_among_nouns, _t_a_feeling_journey, _t_dest_of_feeling,
    _t_nights_adj_noun, _t_three_nouns,
    _t_love_among, _t_love_in_dest,
]

# City-referencing templates — only used when the itinerary has NO activities
# to pull signal from, so names don't mis-attribute features (e.g. "Seminyak
# Paddies" when there are no rice paddies in Seminyak).
CITY_TEMPLATES = [
    _t_two_cities, _t_cities_to, _t_adj_city, _t_city_possessive,
    _t_feeling_in_city, _t_city_noun, _t_adj_city_noun, _t_city_adj_noun,
    _t_noun_of_city, _t_city_and_beyond, _t_verb_in_city, _t_city_love,
]

# Union preserved so LOVE_IDXS is computed against the right set.
TEMPLATES = ACTIVITY_TEMPLATES + CITY_TEMPLATES
LOVE_IDXS_ACTIVITY = {ACTIVITY_TEMPLATES.index(f) for f in (_t_love_among, _t_love_in_dest) if f in ACTIVITY_TEMPLATES}
LOVE_IDXS_CITY = {CITY_TEMPLATES.index(f) for f in (_t_city_love,) if f in CITY_TEMPLATES}

# ─── Public API ───
def generate_name(itinerary, dest="Bali", bank=None, used=None):
    if bank is None: bank = BALI
    if used is None: used = set()

    cities = _normalise_cities(itinerary.get("route") or [])
    nights = itinerary.get("nights") or 0
    activities = itinerary.get("activities") or []
    tags = extract_tags(activities)
    has_activities = bool(activities) and bool(tags)
    base_seed = djb2(f"{itinerary.get('id')}|{','.join(cities)}|{nights}")
    love_allowed = (base_seed % 8 == 0)

    # Choose template pool: activity-driven by default; city-driven only when
    # the itinerary has no activity signal (so names never mis-attribute
    # features to the wrong city).
    if has_activities:
        pool_templates = ACTIVITY_TEMPLATES
        love_idxs = LOVE_IDXS_ACTIVITY
    else:
        pool_templates = CITY_TEMPLATES
        love_idxs = LOVE_IDXS_CITY

    # Try 12× passes through the chosen template pool with seed jitter.
    for attempt in range(len(pool_templates) * 12):
        tmpl_idx = (base_seed + attempt) % len(pool_templates)
        if not love_allowed and tmpl_idx in love_idxs:
            continue
        seed = base_seed + attempt * 7919
        pick = _pick_factory(bank, seed)
        ctx = {"dest": dest, "cities": cities, "nights": nights, "tags": tags, "pick": pick}
        try: name = pool_templates[tmpl_idx](ctx)
        except Exception: name = None
        if not name or len(name) > 24: continue
        if name in used: continue
        used.add(name)
        return name

    # Secondary: if activity pool exhausted, try the city pool (or vice-versa)
    # to expand namespace before falling back to chapter numbering.
    backup = CITY_TEMPLATES if has_activities else ACTIVITY_TEMPLATES
    backup_love = LOVE_IDXS_CITY if has_activities else LOVE_IDXS_ACTIVITY
    for attempt in range(len(backup) * 12):
        tmpl_idx = (base_seed + attempt) % len(backup)
        if not love_allowed and tmpl_idx in backup_love: continue
        seed = base_seed + attempt * 7919 + 13
        pick = _pick_factory(bank, seed)
        ctx = {"dest": dest, "cities": cities, "nights": nights, "tags": tags, "pick": pick}
        try: name = backup[tmpl_idx](ctx)
        except Exception: name = None
        if not name or len(name) > 24: continue
        if name in used: continue
        used.add(name)
        return name

    # Human-tone fallback — never SKU-style.
    # Use dest + chapter numbering, cycle through a bank.
    chapter_pool = ["Chapter", "Edition", "Story", "Tale", "Volume", "Interlude"]
    suffix_idx = 1
    while True:
        word = chapter_pool[(base_seed + suffix_idx) % len(chapter_pool)]
        candidate = f"{dest} {word} {suffix_idx}"
        if candidate not in used and len(candidate) <= 24:
            used.add(candidate)
            return candidate
        suffix_idx += 1
        if suffix_idx > 100000:
            # Truly unreachable — emergency fallback
            fb = f"{dest} #{itinerary.get('id')}"
            used.add(fb)
            return fb


# ─── CLI runner ───
def main():
    import openpyxl
    import sys

    src = "/Users/jaiki96/Downloads/all_destinations_itineraries.xlsx"
    out = "/Users/jaiki96/Downloads/all_destinations_itineraries_named.xlsx"
    sheet_name = sys.argv[1] if len(sys.argv) > 1 else "Bali"
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else 0   # 0 = all

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    new_col = ws.max_column + 1
    if ws.cell(row=1, column=new_col).value is None:
        ws.cell(row=1, column=new_col, value="itinerary_name")

    used = set()
    total = 0
    fallback = 0
    sample = []
    max_row = ws.max_row if not limit else min(ws.max_row, limit + 1)
    for r in range(2, max_row + 1):
        row = [c.value for c in ws[r]]
        try: route = json.loads(row[idx["route"]]) if row[idx["route"]] else []
        except Exception: route = []
        try: activities = json.loads(row[idx["activities"]]) if row[idx["activities"]] else []
        except Exception: activities = []
        it = {"id": row[idx["id"]], "nights": row[idx["nights"]], "route": route, "activities": activities}
        name = generate_name(it, dest=sheet_name, used=used)
        ws.cell(row=r, column=new_col, value=name)
        total += 1
        if re.search(r"\b(Chapter|Edition|Story|Tale|Volume|Interlude)\s+\d+", name):
            fallback += 1
        if r % 2000 == 0:
            print(f"  processed {r - 1}...", flush=True)
        if limit and r <= min(limit + 1, 25):
            sample.append((it["id"], it["nights"], [seg.get("city") for seg in route], name))

    print(f"\nTotal: {total} | Unique: {len(used)} | Fallbacks: {fallback} ({fallback/max(total,1):.1%})")
    if sample:
        print("\nFirst 24 names:")
        for s in sample[:24]: print(f"  id={s[0]:>4} nights={s[1]:>2} cities={s[2]} → {s[3]}")

    if not limit:
        wb.save(out)
        print(f"\nWrote: {out}")


if __name__ == "__main__":
    main()
